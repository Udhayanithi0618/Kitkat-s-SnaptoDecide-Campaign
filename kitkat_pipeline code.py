"""
KitKat Instagram Comment Analysis Pipeline
Modules:
  1. Brand Awareness
  2. Purchase Intention
  3. Customer Engagement
  4. Campaign Performance
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import seaborn as sns
import re
import warnings
from collections import Counter
from datetime import datetime
from textblob import TextBlob
from sklearn.feature_extraction.text import TfidfVectorizer
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

warnings.filterwarnings('ignore')

# ─── Color palette ────────────────────────────────────────────────────────────
RED      = "#C8102E"   # KitKat red
DARK_RED = "#8B0000"
LIGHT_RED= "#FFD6D6"
GREY     = "#F5F5F5"
WHITE    = "#FFFFFF"
DARK     = "#1A1A1A"
GREEN    = "#2ECC71"
ORANGE   = "#E67E22"
BLUE     = "#2980B9"

sns.set_theme(style="whitegrid", palette="deep")
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.titlesize": 13,
    "axes.labelsize": 11,
    "figure.dpi": 150,
})

os.makedirs("/home/claude/charts", exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 – LOAD & PREPROCESS
# ══════════════════════════════════════════════════════════════════════════════
print("▶ Loading data …")
df = pd.read_excel('/mnt/user-data/uploads/Kitkat_Raw_Datasets.xlsx')

# Parse timestamp
df['timestamp'] = pd.to_datetime(pd.to_numeric(df['created_at'], errors='coerce'),
                                  unit='s', errors='coerce')
df = df.dropna(subset=['text'])
df['text'] = df['text'].astype(str).str.strip()

# Remove pure-emoji / single-char noise rows for analysis but keep count
df['text_len']   = df['text'].str.len()
df['is_emoji_only'] = df['text'].str.fullmatch(r'[\U00010000-\U0010ffff\U00002600-\U000027BF\s]+')

def clean_text(t):
    t = re.sub(r'http\S+', '', t)
    t = re.sub(r'@\w+', '', t)
    t = re.sub(r'#\w+', '', t)
    t = re.sub(r'[^\w\s]', ' ', t)
    return t.lower().strip()

df['clean'] = df['text'].apply(clean_text)

# Temporal features
df['date']  = df['timestamp'].dt.date
df['week']  = df['timestamp'].dt.to_period('W')
df['month'] = df['timestamp'].dt.to_period('M')
df['hour']  = df['timestamp'].dt.hour

print(f"   Total comments loaded : {len(df):,}")
print(f"   Date range            : {df['timestamp'].min().date()} → {df['timestamp'].max().date()}")
print(f"   Unique users          : {df['user_id'].nunique():,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 – SENTIMENT ANALYSIS  (TextBlob)
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Analysing sentiment …")

def get_sentiment(text):
    blob = TextBlob(text)
    pol  = blob.sentiment.polarity
    sub  = blob.sentiment.subjectivity
    if pol > 0.15:   label = "Positive"
    elif pol < -0.15: label = "Negative"
    else:             label = "Neutral"
    return pd.Series([pol, sub, label])

df[['polarity','subjectivity','sentiment']] = df['clean'].apply(get_sentiment)

sent_dist = df['sentiment'].value_counts()
print(f"   Positive : {sent_dist.get('Positive',0):,}")
print(f"   Neutral  : {sent_dist.get('Neutral',0):,}")
print(f"   Negative : {sent_dist.get('Negative',0):,}")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 1 – BRAND AWARENESS
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Module 1 – Brand Awareness …")

brand_kw   = ['kitkat','kit kat','kitkat','chocolate','break','wafer','nestle',
              'kitkat bar','kitkat chocolate','candy','snack','crunch']
rival_kw   = ['snickers','twix','kinder','ferrero','mars','bounty','milka',
              'toblerone','cadbury','hershey','oreo','dairy milk']
positive_kw= ['love','great','best','amazing','delicious','yummy','awesome',
              'fantastic','favorite','fav','good','perfect','nice','tasty',
              'want','need','buy','crave']
negative_kw= ['hate','bad','worst','disgusting','terrible','awful','overrated',
              'boring','expensive']

def kw_count(text, kws):
    return sum(1 for k in kws if k in text)

df['brand_mentions']  = df['clean'].apply(lambda x: kw_count(x, brand_kw))
df['rival_mentions']  = df['clean'].apply(lambda x: kw_count(x, rival_kw))
df['positive_brand']  = df['clean'].apply(lambda x: kw_count(x, positive_kw))
df['negative_brand']  = df['clean'].apply(lambda x: kw_count(x, negative_kw))

total_comments    = len(df)
brand_aware_cnt   = (df['brand_mentions'] > 0).sum()
brand_awareness_pct = brand_aware_cnt / total_comments * 100

# Top words
all_words = ' '.join(df['clean']).split()
stop = set(['the','a','an','is','it','in','of','and','to','for','i','me',
            'my','you','your','we','this','that','are','was','on','at','be',
            'or','do','not','with','have','has','but','as','from','by',
            'he','she','they','so','if','its','we','our','their','can',
            'will','just','all','also','like','no','yes','more','up','one'])
filtered = [w for w in all_words if w not in stop and len(w) > 2 and w.isalpha()]
word_freq = Counter(filtered).most_common(30)

# Rival co-mention
rival_co = {r: (df['clean'].str.contains(r)).sum() for r in rival_kw}
rival_co = {k:v for k,v in sorted(rival_co.items(), key=lambda x: -x[1]) if v > 0}

print(f"   Brand awareness rate  : {brand_awareness_pct:.1f}%")
print(f"   Brand-aware comments  : {brand_aware_cnt:,}")
print(f"   Top rival co-mentions : {list(rival_co.items())[:5]}")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 2 – PURCHASE INTENTION
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Module 2 – Purchase Intention …")

buy_kw   = ['buy','purchase','order','shop','get','want','need','craving',
            'crave','will buy','gonna buy','must buy','gonna get','need one',
            'add to cart','checkout','store']
neg_int  = ['never buy','won\'t buy','wont buy','don\'t buy','not buying',
            'stopped buying','boycott','avoid','skip']

def classify_intent(text):
    if any(k in text for k in neg_int):  return "Low Intent (Negative)"
    if any(k in text for k in buy_kw):   return "High Intent"
    if any(k in ['consider','maybe','might','perhaps','thinking'] for k in text.split()):
        return "Medium Intent"
    return "Low Intent"

df['purchase_intent'] = df['clean'].apply(classify_intent)
intent_dist = df['purchase_intent'].value_counts()

# Purchase intent by sentiment
intent_sent = pd.crosstab(df['purchase_intent'], df['sentiment'])

print(f"   High Intent comments  : {intent_dist.get('High Intent',0):,}")
print(f"   Medium Intent         : {intent_dist.get('Medium Intent',0):,}")
print(f"   Low Intent            : {intent_dist.get('Low Intent',0):,}")
print(f"   Low Intent (Neg)      : {intent_dist.get('Low Intent (Negative)',0):,}")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 3 – CUSTOMER ENGAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Module 3 – Customer Engagement …")

# Engagement metrics
df['has_mention']  = df['text'].str.contains(r'@\w+', na=False)
df['has_hashtag']  = df['text'].str.contains(r'#\w+', na=False)
df['has_emoji']    = df['text'].str.contains(
    r'[\U0001F300-\U0001FFFF\U00002600-\U000027BF]', na=False)
df['is_question']  = df['text'].str.contains(r'\?', na=False)
df['reply_tag']    = df['has_mention']

# Engagement score (composite)
df['engagement_score'] = (
    df['text_len'].clip(0,200)/200 * 30 +
    df['has_mention'].astype(int)   * 25 +
    df['has_hashtag'].astype(int)   * 15 +
    df['has_emoji'].astype(int)     * 15 +
    df['is_question'].astype(int)   * 15
)

# Daily volume
daily = df.groupby('date').size().reset_index(name='comments')
daily['date'] = pd.to_datetime(daily['date'])
daily = daily.sort_values('date')
daily['rolling_7'] = daily['comments'].rolling(7, min_periods=1).mean()

# Hourly pattern
hourly = df.groupby('hour').size().reset_index(name='comments')

# Power users
power_users = (df.groupby('username')['comment_id']
                 .count()
                 .reset_index(name='comment_count')
                 .sort_values('comment_count', ascending=False)
                 .head(15))

avg_eng   = df['engagement_score'].mean()
pct_reply = df['has_mention'].mean() * 100
pct_emoji = df['has_emoji'].mean()   * 100

print(f"   Avg engagement score  : {avg_eng:.1f}/100")
print(f"   Comments with @mention: {pct_reply:.1f}%")
print(f"   Comments with emoji   : {pct_emoji:.1f}%")


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 4 – CAMPAIGN PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Module 4 – Campaign Performance …")

weekly = df.groupby('week').agg(
    comments      = ('comment_id','count'),
    avg_polarity  = ('polarity','mean'),
    avg_eng       = ('engagement_score','mean'),
    high_intent   = ('purchase_intent', lambda x: (x=='High Intent').sum()),
    positive      = ('sentiment', lambda x: (x=='Positive').sum()),
    negative      = ('sentiment', lambda x: (x=='Negative').sum()),
).reset_index()
weekly['week_str']     = weekly['week'].astype(str)
weekly['intent_rate']  = weekly['high_intent'] / weekly['comments'] * 100
weekly['pos_rate']     = weekly['positive']    / weekly['comments'] * 100
weekly['neg_rate']     = weekly['negative']    / weekly['comments'] * 100
weekly['net_sentiment'] = weekly['pos_rate'] - weekly['neg_rate']

# Peak week
peak_week = weekly.loc[weekly['comments'].idxmax(), 'week_str']
peak_vol  = weekly['comments'].max()

# KPIs
kpis = {
    "Total Comments"         : f"{total_comments:,}",
    "Unique Users"           : f"{df['user_id'].nunique():,}",
    "Brand Awareness Rate"   : f"{brand_awareness_pct:.1f}%",
    "Avg Polarity"           : f"{df['polarity'].mean():.3f}",
    "High Purchase Intent"   : f"{intent_dist.get('High Intent',0):,}",
    "Intent Rate"            : f"{intent_dist.get('High Intent',0)/total_comments*100:.1f}%",
    "Avg Engagement Score"   : f"{avg_eng:.1f}/100",
    "% Positive Comments"    : f"{sent_dist.get('Positive',0)/total_comments*100:.1f}%",
    "% Negative Comments"    : f"{sent_dist.get('Negative',0)/total_comments*100:.1f}%",
    "Peak Activity Week"     : peak_week,
    "Peak Week Volume"       : f"{peak_vol:,}",
}

print(f"   Peak activity week    : {peak_week} ({peak_vol:,} comments)")


# ══════════════════════════════════════════════════════════════════════════════
# CHARTS
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Generating charts …")

# ── Chart 1: Sentiment Distribution ──────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
fig.suptitle("KitKat Instagram – Sentiment Analysis", fontsize=15, fontweight='bold', color=DARK)

colors_sent = {'Positive': GREEN, 'Neutral': BLUE, 'Negative': RED}
ax = axes[0]
vals  = [sent_dist.get(s, 0) for s in ['Positive','Neutral','Negative']]
bars  = ax.bar(['Positive','Neutral','Negative'], vals,
               color=[GREEN, BLUE, RED], edgecolor='white', linewidth=1.5,
               width=0.5)
for b, v in zip(bars, vals):
    ax.text(b.get_x()+b.get_width()/2, b.get_height()+100, f'{v:,}',
            ha='center', va='bottom', fontsize=10, fontweight='bold')
ax.set_title("Sentiment Distribution", fontweight='bold')
ax.set_ylabel("Number of Comments")
ax.spines[['top','right']].set_visible(False)

ax2 = axes[1]
explode = (0.05, 0.02, 0.05)
wedges, texts, autotexts = ax2.pie(vals,
    labels=['Positive','Neutral','Negative'],
    colors=[GREEN, BLUE, RED],
    autopct='%1.1f%%', startangle=140, explode=explode,
    textprops={'fontsize': 10})
for at in autotexts: at.set_fontweight('bold')
ax2.set_title("Sentiment Share", fontweight='bold')

plt.tight_layout()
plt.savefig("/home/claude/charts/01_sentiment.png", bbox_inches='tight')
plt.close()

# ── Chart 2: Brand Awareness ──────────────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle("KitKat – Brand Awareness Analysis", fontsize=15, fontweight='bold', color=DARK)

# Word frequency bar
wf_labels = [w[0] for w in word_freq[:15]]
wf_values = [w[1] for w in word_freq[:15]]
ax = axes[0]
cmap = plt.cm.Reds(np.linspace(0.4, 0.9, 15))[::-1]
hbars = ax.barh(wf_labels[::-1], wf_values[::-1], color=cmap[::-1],
                edgecolor='white')
ax.set_title("Top 15 Keywords", fontweight='bold')
ax.set_xlabel("Frequency")
for bar, val in zip(hbars, wf_values[::-1]):
    ax.text(bar.get_width()+5, bar.get_y()+bar.get_height()/2,
            f'{val:,}', va='center', fontsize=8)
ax.spines[['top','right']].set_visible(False)

# Rival co-mentions
ax2 = axes[1]
rv_names = list(rival_co.keys())[:8]
rv_vals  = [rival_co[k] for k in rv_names]
ax2.bar(rv_names, rv_vals, color=DARK_RED, alpha=0.85,
        edgecolor='white', linewidth=1)
ax2.set_title("Competitor Co-Mentions", fontweight='bold')
ax2.set_ylabel("Mentions")
ax2.set_xticklabels(rv_names, rotation=30, ha='right')
ax2.spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig("/home/claude/charts/02_brand_awareness.png", bbox_inches='tight')
plt.close()

# ── Chart 3: Purchase Intention ───────────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
fig.suptitle("KitKat – Purchase Intention Analysis", fontsize=15, fontweight='bold', color=DARK)

intent_colors = {'High Intent': GREEN, 'Medium Intent': ORANGE,
                 'Low Intent': BLUE, 'Low Intent (Negative)': RED}
ax = axes[0]
i_cats = intent_dist.index.tolist()
i_vals = intent_dist.values.tolist()
i_cols = [intent_colors.get(c, GREY) for c in i_cats]
wedges2, texts2, autos2 = ax.pie(i_vals, labels=i_cats, colors=i_cols,
    autopct='%1.1f%%', startangle=90,
    textprops={'fontsize': 9})
for at in autos2: at.set_fontweight('bold')
ax.set_title("Intent Distribution", fontweight='bold')

# Stacked bar: intent × sentiment
ax2 = axes[1]
intent_sent2 = pd.crosstab(df['purchase_intent'], df['sentiment'])
intent_sent2 = intent_sent2.reindex(index=['High Intent','Medium Intent',
                                            'Low Intent','Low Intent (Negative)'])
bottom = np.zeros(len(intent_sent2))
sent_colors = {'Positive': GREEN, 'Neutral': BLUE, 'Negative': RED}
for col in intent_sent2.columns:
    vals_s = intent_sent2[col].fillna(0).values
    ax2.bar(intent_sent2.index, vals_s, bottom=bottom,
            color=sent_colors.get(col, GREY), label=col,
            edgecolor='white', linewidth=0.8)
    bottom += vals_s
ax2.set_title("Intent × Sentiment", fontweight='bold')
ax2.set_ylabel("Comments")
ax2.set_xticklabels(intent_sent2.index, rotation=20, ha='right', fontsize=8)
ax2.legend(fontsize=8)
ax2.spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig("/home/claude/charts/03_purchase_intent.png", bbox_inches='tight')
plt.close()

# ── Chart 4: Customer Engagement ─────────────────────────────────────────────
fig, axes = plt.subplots(2, 2, figsize=(14, 9))
fig.suptitle("KitKat – Customer Engagement Analysis", fontsize=15, fontweight='bold', color=DARK)

# Hourly pattern
ax = axes[0,0]
ax.plot(hourly['hour'], hourly['comments'], color=RED, linewidth=2.5,
        marker='o', markersize=4)
ax.fill_between(hourly['hour'], hourly['comments'], alpha=0.2, color=RED)
ax.set_title("Comment Activity by Hour (UTC)", fontweight='bold')
ax.set_xlabel("Hour of Day")
ax.set_ylabel("Comments")
ax.set_xticks(range(0, 24, 2))
ax.spines[['top','right']].set_visible(False)

# Engagement score distribution
ax2 = axes[0,1]
ax2.hist(df['engagement_score'], bins=30, color=RED, edgecolor='white',
         alpha=0.85)
ax2.axvline(avg_eng, color=DARK, linestyle='--', linewidth=2,
            label=f'Mean = {avg_eng:.1f}')
ax2.set_title("Engagement Score Distribution", fontweight='bold')
ax2.set_xlabel("Engagement Score")
ax2.set_ylabel("Frequency")
ax2.legend(fontsize=9)
ax2.spines[['top','right']].set_visible(False)

# Feature flags
ax3 = axes[1,0]
flags = {
    'Has @Mention' : df['has_mention'].mean()*100,
    'Has Emoji'    : df['has_emoji'].mean()*100,
    'Is Question'  : df['is_question'].mean()*100,
    'Has Hashtag'  : df['has_hashtag'].mean()*100,
}
flag_colors = [RED, ORANGE, BLUE, GREEN]
bars3 = ax3.bar(list(flags.keys()), list(flags.values()),
                color=flag_colors, edgecolor='white', linewidth=1.2,
                width=0.5)
for b, v in zip(bars3, flags.values()):
    ax3.text(b.get_x()+b.get_width()/2, b.get_height()+0.3,
             f'{v:.1f}%', ha='center', fontsize=9, fontweight='bold')
ax3.set_title("Engagement Feature Breakdown", fontweight='bold')
ax3.set_ylabel("% of Comments")
ax3.set_ylim(0, max(flags.values())*1.2)
ax3.spines[['top','right']].set_visible(False)

# Power users
ax4 = axes[1,1]
top10 = power_users.head(10)
ax4.barh(top10['username'][::-1], top10['comment_count'][::-1],
         color=RED, alpha=0.85, edgecolor='white')
ax4.set_title("Top 10 Most Active Users", fontweight='bold')
ax4.set_xlabel("Comments Posted")
ax4.spines[['top','right']].set_visible(False)

plt.tight_layout()
plt.savefig("/home/claude/charts/04_engagement.png", bbox_inches='tight')
plt.close()

# ── Chart 5: Campaign Performance ────────────────────────────────────────────
fig = plt.figure(figsize=(15, 10))
fig.suptitle("KitKat – Campaign Performance Dashboard", fontsize=16,
             fontweight='bold', color=DARK)
gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.4, wspace=0.35)

# 5a: Daily volume + 7-day rolling
ax1 = fig.add_subplot(gs[0, :])
ax1.bar(daily['date'], daily['comments'], color=LIGHT_RED,
        edgecolor='none', alpha=0.9, label='Daily Comments')
ax1.plot(daily['date'], daily['rolling_7'], color=RED, linewidth=2.5,
         label='7-day Rolling Avg')
ax1.set_title("Daily Comment Volume & 7-Day Trend", fontweight='bold')
ax1.set_ylabel("Comments")
ax1.legend(fontsize=9)
ax1.spines[['top','right']].set_visible(False)

# 5b: Weekly sentiment net
ax2 = fig.add_subplot(gs[1, 0])
colors_net = [GREEN if v >= 0 else RED for v in weekly['net_sentiment']]
ax2.bar(range(len(weekly)), weekly['net_sentiment'], color=colors_net,
        edgecolor='white', linewidth=0.8)
ax2.axhline(0, color=DARK, linewidth=0.8)
ax2.set_title("Weekly Net Sentiment Score\n(Positive% − Negative%)", fontweight='bold')
ax2.set_xlabel("Week Index")
ax2.set_ylabel("Net Sentiment (%)")
ax2.spines[['top','right']].set_visible(False)

# 5c: Weekly purchase intent rate
ax3 = fig.add_subplot(gs[1, 1])
ax3.plot(range(len(weekly)), weekly['intent_rate'], color=GREEN,
         linewidth=2.5, marker='o', markersize=5)
ax3.fill_between(range(len(weekly)), weekly['intent_rate'],
                 alpha=0.15, color=GREEN)
ax3.set_title("Weekly High Purchase Intent Rate (%)", fontweight='bold')
ax3.set_xlabel("Week Index")
ax3.set_ylabel("Intent Rate (%)")
ax3.spines[['top','right']].set_visible(False)

plt.savefig("/home/claude/charts/05_campaign.png", bbox_inches='tight')
plt.close()

print("   Charts saved ✓")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════
print("\n▶ Building Excel report …")

wb = Workbook()
wb.remove(wb.active)

# ─── Styles ───────────────────────────────────────────────────────────────────
def hdr(ws, row, col, val, bold=True, size=11, bg=None, fg='FFFFFF',
        align='center', wrap=False, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=size, color=fg,
                     name='Arial')
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=wrap)
    if border:
        thin = Side(style='thin', color='DDDDDD')
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return cell

def data_cell(ws, row, col, val, align='center', number_format=None,
               bold=False, bg=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name='Arial', size=10, bold=bold,
                     color='1A1A1A')
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=False)
    thin = Side(style='thin', color='EEEEEE')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

RED_HEX  = 'C8102E'
DRED_HEX = '8B0000'
GRN_HEX  = '2ECC71'
BLU_HEX  = '2980B9'
ORG_HEX  = 'E67E22'
LGREY    = 'F5F5F5'
MGREY    = 'E0E0E0'
WHITE_H  = 'FFFFFF'


# ─── SHEET 1: KPI Dashboard ───────────────────────────────────────────────────
ws1 = wb.create_sheet("📊 KPI Dashboard")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A4'

# Title banner
ws1.merge_cells('A1:H1')
title_cell = ws1['A1']
title_cell.value = "🍫  KITKAT INSTAGRAM ANALYTICS PIPELINE  |  Brand Intelligence Report"
title_cell.font  = Font(bold=True, size=16, color=WHITE_H, name='Arial')
title_cell.fill  = PatternFill('solid', fgColor=DRED_HEX)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:H2')
sub = ws1['A2']
sub.value = f"Generated: {datetime.now().strftime('%d %B %Y')}  |  Dataset: KitKat Instagram Comments  |  Total Records: {total_comments:,}"
sub.font  = Font(italic=True, size=10, color='666666', name='Arial')
sub.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 20

ws1.row_dimensions[3].height = 10  # spacer

# KPI header
hdr(ws1, 4, 1, "KEY PERFORMANCE INDICATORS", size=12, bg=RED_HEX, fg=WHITE_H,
    align='center')
ws1.merge_cells('A4:H4')
ws1.row_dimensions[4].height = 28

# KPI labels
kpi_rows = [
    ("📌 Total Comments Analysed",    f"{total_comments:,}",            "Volume of engagement captured"),
    ("👥 Unique Users",               f"{df['user_id'].nunique():,}",   "Individual commenters"),
    ("🏷️ Brand Awareness Rate",       f"{brand_awareness_pct:.1f}%",    "% comments with brand keywords"),
    ("💚 Avg Sentiment Polarity",     f"{df['polarity'].mean():.3f}",   "Scale: -1 (neg) to +1 (pos)"),
    ("🛒 High Purchase Intent",       f"{intent_dist.get('High Intent',0):,}", "Comments showing buy intent"),
    ("📈 Intent Conversion Rate",     f"{intent_dist.get('High Intent',0)/total_comments*100:.1f}%",
     "% high-intent out of total"),
    ("⚡ Avg Engagement Score",       f"{avg_eng:.1f}/100",             "Composite engagement metric"),
    ("😊 % Positive Comments",        f"{sent_dist.get('Positive',0)/total_comments*100:.1f}%",
     "Positive sentiment share"),
    ("😟 % Negative Comments",        f"{sent_dist.get('Negative',0)/total_comments*100:.1f}%",
     "Negative sentiment share"),
    ("📅 Peak Activity Week",         peak_week,                        f"Highest volume: {peak_vol:,} comments"),
    ("🔁 Comments with @Mention",     f"{pct_reply:.1f}%",             "Conversational engagement"),
    ("😂 Comments with Emoji",        f"{pct_emoji:.1f}%",             "Emoji-based expression"),
]

for i, (label, value, note) in enumerate(kpi_rows):
    r = 5 + i
    ws1.row_dimensions[r].height = 22
    bg = LGREY if i % 2 == 0 else WHITE_H
    data_cell(ws1, r, 1, label, align='left', bg=bg, bold=True)
    data_cell(ws1, r, 2, value, bg=bg, bold=True)
    data_cell(ws1, r, 3, note, align='left', bg=bg)
    ws1.merge_cells(f'C{r}:H{r}')

for col, w in [(1,40),(2,20),(3,50)]:
    set_col_width(ws1, col, w)


# ─── SHEET 2: Brand Awareness ─────────────────────────────────────────────────
ws2 = wb.create_sheet("🏷️ Brand Awareness")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A4'

ws2.merge_cells('A1:F1')
c = ws2['A1']
c.value = "BRAND AWARENESS ANALYSIS"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=RED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

# Top keywords table
hdr(ws2, 3, 1, "#",    bg=DRED_HEX, size=10)
hdr(ws2, 3, 2, "Keyword", bg=DRED_HEX, size=10, align='left')
hdr(ws2, 3, 3, "Frequency", bg=DRED_HEX, size=10)
hdr(ws2, 3, 4, "% of Comments", bg=DRED_HEX, size=10)
hdr(ws2, 3, 5, "Rank", bg=DRED_HEX, size=10)

for i, (word, freq) in enumerate(word_freq[:20]):
    r   = 4 + i
    bg  = LGREY if i % 2 == 0 else WHITE_H
    pct = freq / total_comments * 100
    data_cell(ws2, r, 1, i+1, bg=bg)
    data_cell(ws2, r, 2, word.title(), align='left', bg=bg)
    data_cell(ws2, r, 3, freq, bg=bg)
    data_cell(ws2, r, 4, round(pct, 2), bg=bg, number_format='0.00%')
    data_cell(ws2, r, 5, "🔴" if i < 5 else "🟡" if i < 10 else "⚪", bg=bg)
    ws2.row_dimensions[r].height = 18

# Competitor section
r_start = 26
hdr(ws2, r_start, 1, "COMPETITOR CO-MENTIONS", bg=DRED_HEX,
    size=11, align='left')
ws2.merge_cells(f'A{r_start}:F{r_start}')
hdr(ws2, r_start+1, 1, "Brand",     bg=RED_HEX, size=10)
hdr(ws2, r_start+1, 2, "Mentions",  bg=RED_HEX, size=10)
hdr(ws2, r_start+1, 3, "% of Total", bg=RED_HEX, size=10)

for j, (brand, cnt) in enumerate(rival_co.items()):
    r   = r_start + 2 + j
    bg  = LGREY if j % 2 == 0 else WHITE_H
    pct = cnt / total_comments * 100
    data_cell(ws2, r, 1, brand.title(), align='left', bg=bg)
    data_cell(ws2, r, 2, cnt, bg=bg)
    data_cell(ws2, r, 3, round(pct, 3), bg=bg, number_format='0.000%')
    ws2.row_dimensions[r].height = 18

for col, w in [(1,6),(2,20),(3,15),(4,18),(5,10)]:
    set_col_width(ws2, col, w)


# ─── SHEET 3: Sentiment Analysis ─────────────────────────────────────────────
ws3 = wb.create_sheet("😊 Sentiment")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'A4'

ws3.merge_cells('A1:G1')
c = ws3['A1']
c.value = "SENTIMENT ANALYSIS"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=RED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

# Sentiment summary
for col, label in [(1,'Metric'),(2,'Value'),(3,'Share')]:
    hdr(ws3, 3, col, label, bg=DRED_HEX, size=10)

rows3 = [
    ("Total Comments",  total_comments, "100.0%"),
    ("✅ Positive",      sent_dist.get('Positive',0),
     f"{sent_dist.get('Positive',0)/total_comments*100:.1f}%"),
    ("⬜ Neutral",       sent_dist.get('Neutral',0),
     f"{sent_dist.get('Neutral',0)/total_comments*100:.1f}%"),
    ("❌ Negative",      sent_dist.get('Negative',0),
     f"{sent_dist.get('Negative',0)/total_comments*100:.1f}%"),
    ("Avg Polarity",    round(df['polarity'].mean(),4), "(-1 neg → +1 pos)"),
    ("Avg Subjectivity",round(df['subjectivity'].mean(),4), "(0 obj → 1 subj)"),
]
for i, (m, v, s) in enumerate(rows3):
    r  = 4 + i
    bg = LGREY if i % 2 == 0 else WHITE_H
    data_cell(ws3, r, 1, m, align='left', bg=bg, bold=True)
    data_cell(ws3, r, 2, v, bg=bg)
    data_cell(ws3, r, 3, s, bg=bg)
    ws3.row_dimensions[r].height = 20

# Weekly sentiment table
r_s = 12
hdr(ws3, r_s, 1, "WEEKLY SENTIMENT TREND", bg=DRED_HEX,
    size=11, align='left')
ws3.merge_cells(f'A{r_s}:G{r_s}')

headers_w = ['Week','Comments','Avg Polarity','Positive%','Neutral%',
             'Negative%','Net Sentiment']
for ci, h in enumerate(headers_w, 1):
    hdr(ws3, r_s+1, ci, h, bg=RED_HEX, size=9)

for j, row in weekly.iterrows():
    r  = r_s + 2 + j
    bg = LGREY if j % 2 == 0 else WHITE_H
    data_cell(ws3, r, 1, row['week_str'], bg=bg)
    data_cell(ws3, r, 2, int(row['comments']), bg=bg)
    data_cell(ws3, r, 3, round(row['avg_polarity'],3), bg=bg)
    data_cell(ws3, r, 4, round(row['pos_rate'],1), bg=bg)
    data_cell(ws3, r, 5, round(100-row['pos_rate']-row['neg_rate'],1), bg=bg)
    data_cell(ws3, r, 6, round(row['neg_rate'],1), bg=bg)
    net = round(row['net_sentiment'],1)
    net_bg = '90EE90' if net >= 0 else 'FFB3B3'
    data_cell(ws3, r, 7, net, bg=net_bg, bold=True)
    ws3.row_dimensions[r].height = 16

for col, w in [(1,18),(2,14),(3,14),(4,13),(5,13),(6,13),(7,15)]:
    set_col_width(ws3, col, w)


# ─── SHEET 4: Purchase Intention ─────────────────────────────────────────────
ws4 = wb.create_sheet("🛒 Purchase Intention")
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
c = ws4['A1']
c.value = "PURCHASE INTENTION ANALYSIS"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=RED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 35

intent_color_map = {
    'High Intent'            : '2ECC71',
    'Medium Intent'          : 'E67E22',
    'Low Intent'             : '2980B9',
    'Low Intent (Negative)'  : 'C8102E',
}

hdr(ws4, 3, 1, "Intent Level",   bg=DRED_HEX)
hdr(ws4, 3, 2, "Comments",       bg=DRED_HEX)
hdr(ws4, 3, 3, "Share %",        bg=DRED_HEX)
hdr(ws4, 3, 4, "Positive",       bg=DRED_HEX)
hdr(ws4, 3, 5, "Neutral",        bg=DRED_HEX)
hdr(ws4, 3, 6, "Negative",       bg=DRED_HEX)
hdr(ws4, 3, 7, "Avg Polarity",   bg=DRED_HEX)

intent_order = ['High Intent','Medium Intent','Low Intent','Low Intent (Negative)']
for i, intent in enumerate(intent_order):
    r    = 4 + i
    sub_ = df[df['purchase_intent'] == intent]
    cnt  = len(sub_)
    pct  = cnt / total_comments * 100
    pos  = (sub_['sentiment'] == 'Positive').sum()
    neu  = (sub_['sentiment'] == 'Neutral').sum()
    neg  = (sub_['sentiment'] == 'Negative').sum()
    pol  = round(sub_['polarity'].mean(), 3) if cnt > 0 else 0
    ibg  = intent_color_map.get(intent, LGREY)

    data_cell(ws4, r, 1, intent, align='left',
              bg=ibg, bold=True)
    ws4.cell(r, 1).font = Font(bold=True, color=WHITE_H, name='Arial', size=10)
    for ci, v in enumerate([cnt, round(pct,1), pos, neu, neg, pol], 2):
        data_cell(ws4, r, ci, v, bg=LGREY if i%2==0 else WHITE_H)
    ws4.row_dimensions[r].height = 22

# Sample high-intent comments
r_si = 10
ws4.merge_cells(f'A{r_si}:G{r_si}')
hdr(ws4, r_si, 1, "SAMPLE HIGH PURCHASE INTENT COMMENTS",
    bg=DRED_HEX, align='left')

high = df[df['purchase_intent']=='High Intent'][['text','polarity']].head(20)
for ci, col in enumerate(['Comment Text','Polarity'], 1):
    hdr(ws4, r_si+1, ci, col, bg=RED_HEX, size=9,
        align='left' if ci==1 else 'center')

for j, (_, row) in enumerate(high.iterrows()):
    r  = r_si + 2 + j
    bg = LGREY if j%2==0 else WHITE_H
    data_cell(ws4, r, 1, str(row['text'])[:120], align='left', bg=bg)
    pol_bg = '90EE90' if row['polarity'] > 0 else 'FFB3B3' if row['polarity'] < 0 else bg
    data_cell(ws4, r, 2, round(row['polarity'],3), bg=pol_bg)
    ws4.row_dimensions[r].height = 18

ws4.merge_cells(f'B{r_si+1}:G{r_si+1}')  # fix header merge
for col, w in [(1,50),(2,12),(3,12),(4,12),(5,12),(6,12),(7,14)]:
    set_col_width(ws4, col, w)


# ─── SHEET 5: Engagement Analysis ────────────────────────────────────────────
ws5 = wb.create_sheet("⚡ Engagement")
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:G1')
c = ws5['A1']
c.value = "CUSTOMER ENGAGEMENT ANALYSIS"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=RED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 35

# Engagement feature summary
hdr(ws5, 3, 1, "Engagement Feature", bg=DRED_HEX, align='left')
hdr(ws5, 3, 2, "Count", bg=DRED_HEX)
hdr(ws5, 3, 3, "% of Total", bg=DRED_HEX)

eng_features = [
    ("@Mention (Conversational)", df['has_mention'].sum()),
    ("Emoji Usage",               df['has_emoji'].sum()),
    ("Question Marks (Enquiry)",  df['is_question'].sum()),
    ("Hashtag Usage",             df['has_hashtag'].sum()),
]
for i, (feat, cnt) in enumerate(eng_features):
    r   = 4 + i
    bg  = LGREY if i%2==0 else WHITE_H
    pct = cnt/total_comments*100
    data_cell(ws5, r, 1, feat, align='left', bg=bg, bold=True)
    data_cell(ws5, r, 2, int(cnt), bg=bg)
    data_cell(ws5, r, 3, round(pct,1), bg=bg)
    ws5.row_dimensions[r].height = 20

# Hourly data
r_h = 10
hdr(ws5, r_h, 1, "HOURLY ENGAGEMENT PATTERN", bg=DRED_HEX,
    align='left', size=11)
ws5.merge_cells(f'A{r_h}:G{r_h}')
hdr(ws5, r_h+1, 1, "Hour (UTC)", bg=RED_HEX)
hdr(ws5, r_h+1, 2, "Comments", bg=RED_HEX)
hdr(ws5, r_h+1, 3, "% of Daily", bg=RED_HEX)

for j, row_h in hourly.iterrows():
    r   = r_h + 2 + j
    bg  = LGREY if j%2==0 else WHITE_H
    pct = row_h['comments']/total_comments*100
    data_cell(ws5, r, 1, int(row_h['hour']), bg=bg)
    data_cell(ws5, r, 2, int(row_h['comments']), bg=bg)
    data_cell(ws5, r, 3, round(pct,2), bg=bg)
    ws5.row_dimensions[r].height = 16

# Power users
r_pu = 36
hdr(ws5, r_pu, 1, "TOP 15 POWER USERS", bg=DRED_HEX, align='left', size=11)
ws5.merge_cells(f'A{r_pu}:G{r_pu}')
hdr(ws5, r_pu+1, 1, "Rank", bg=RED_HEX)
hdr(ws5, r_pu+1, 2, "Username", bg=RED_HEX, align='left')
hdr(ws5, r_pu+1, 3, "Comments", bg=RED_HEX)

for k, (_, pu) in enumerate(power_users.iterrows()):
    r   = r_pu + 2 + k
    bg  = 'FFD700' if k == 0 else ('C0C0C0' if k == 1 else
          ('CD7F32' if k == 2 else (LGREY if k%2==0 else WHITE_H)))
    data_cell(ws5, r, 1, k+1, bg=bg, bold=(k<3))
    data_cell(ws5, r, 2, pu['username'], align='left', bg=bg, bold=(k<3))
    data_cell(ws5, r, 3, int(pu['comment_count']), bg=bg, bold=(k<3))
    ws5.row_dimensions[r].height = 18

for col, w in [(1,22),(2,28),(3,16)]:
    set_col_width(ws5, col, w)


# ─── SHEET 6: Campaign Performance ───────────────────────────────────────────
ws6 = wb.create_sheet("📈 Campaign Performance")
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = 'A4'

ws6.merge_cells('A1:I1')
c = ws6['A1']
c.value = "CAMPAIGN PERFORMANCE – WEEKLY BREAKDOWN"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=RED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 35

headers6 = ['Week','Comments','Avg Polarity','Avg Eng. Score',
            'High Intent','Intent Rate%','Positive%','Negative%',
            'Net Sentiment']
for ci, h in enumerate(headers6, 1):
    hdr(ws6, 3, ci, h, bg=DRED_HEX, size=9)
    ws6.row_dimensions[3].height = 22

for j, row in weekly.iterrows():
    r  = 4 + j
    bg = LGREY if j%2==0 else WHITE_H
    vals6 = [
        row['week_str'],
        int(row['comments']),
        round(row['avg_polarity'],3),
        round(row['avg_eng'],1),
        int(row['high_intent']),
        round(row['intent_rate'],1),
        round(row['pos_rate'],1),
        round(row['neg_rate'],1),
        round(row['net_sentiment'],1),
    ]
    for ci, v in enumerate(vals6, 1):
        cell = data_cell(ws6, r, ci, v, bg=bg)
        if ci == 9:  # Net sentiment coloring
            cell.fill = PatternFill('solid',
                fgColor='90EE90' if isinstance(v, float) and v >= 0 else 'FFB3B3')
            cell.font = Font(bold=True, name='Arial', size=10,
                             color='1A1A1A')
    ws6.row_dimensions[r].height = 16

for col, w in [(1,22),(2,13),(3,14),(4,16),(5,13),(6,13),(7,13),(8,13),(9,15)]:
    set_col_width(ws6, col, w)


# ─── SHEET 7: Raw Data Sample ─────────────────────────────────────────────────
ws7 = wb.create_sheet("📋 Raw Data Sample")
ws7.sheet_view.showGridLines = False

sample = df[['comment_id','timestamp','username','text','sentiment',
             'polarity','subjectivity','purchase_intent',
             'engagement_score','has_mention','has_emoji']].head(500)

headers7 = ['Comment ID','Timestamp','Username','Comment Text',
            'Sentiment','Polarity','Subjectivity','Purchase Intent',
            'Eng. Score','Has @Mention','Has Emoji']
for ci, h in enumerate(headers7, 1):
    hdr(ws7, 1, ci, h, bg=DRED_HEX, size=9)

for i, (_, row) in enumerate(sample.iterrows()):
    r  = 2 + i
    bg = LGREY if i%2==0 else WHITE_H
    sent_bg = {'Positive':'90EE90','Negative':'FFB3B3',
               'Neutral':LGREY}.get(str(row['sentiment']), LGREY)
    vals7 = [
        str(row['comment_id']),
        str(row['timestamp'])[:16],
        str(row['username']),
        str(row['text'])[:100],
        str(row['sentiment']),
        round(float(row['polarity']),3),
        round(float(row['subjectivity']),3),
        str(row['purchase_intent']),
        round(float(row['engagement_score']),1),
        "Yes" if row['has_mention'] else "No",
        "Yes" if row['has_emoji'] else "No",
    ]
    for ci, v in enumerate(vals7, 1):
        cell = data_cell(ws7, r, ci, v,
                         align='left' if ci in [3,4,5,8] else 'center',
                         bg=sent_bg if ci==5 else bg)
    ws7.row_dimensions[r].height = 16

for col, w in [(1,20),(2,18),(3,18),(4,50),(5,12),(6,10),(7,13),(8,22),(9,12),(10,13),(11,11)]:
    set_col_width(ws7, col, w)
ws7.freeze_panes = 'A2'


# ─── SHEET 8: Methodology ────────────────────────────────────────────────────
ws8 = wb.create_sheet("📖 Methodology")
ws8.sheet_view.showGridLines = False

ws8.merge_cells('A1:E1')
c = ws8['A1']
c.value = "PIPELINE METHODOLOGY & DEFINITIONS"
c.font  = Font(bold=True, size=14, color=WHITE_H, name='Arial')
c.fill  = PatternFill('solid', fgColor=DRED_HEX)
c.alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 35

methods = [
    ("MODULE", "TECHNIQUE", "DESCRIPTION", "TOOLS", "OUTPUT"),
    ("1. Brand Awareness", "Keyword Frequency Analysis",
     "Count of brand/competitor keywords across comments. Brand awareness rate = brand-mentions / total.",
     "Python Counter, Regex", "Awareness %, Word Freq, Rival Co-mentions"),
    ("2. Sentiment Analysis", "TextBlob Lexicon-based NLP",
     "Polarity (-1 to +1) and Subjectivity (0-1). Positive > 0.15, Negative < -0.15, else Neutral.",
     "TextBlob", "Sentiment label, Polarity score, Subjectivity"),
    ("3. Purchase Intention", "Rule-based Intent Classifier",
     "High: contains buy/purchase/want/need/crave. Neg: contains boycott/won't buy. Medium: uncertainty words.",
     "Regex pattern matching", "Intent label (4 classes)"),
    ("4. Customer Engagement", "Composite Engagement Score",
     "Score = (text_len/200×30) + @mention×25 + hashtag×15 + emoji×15 + question×15. Max = 100.",
     "Pandas, NumPy", "Engagement score, Hourly patterns, Power users"),
    ("5. Campaign Performance", "Weekly Time-series Aggregation",
     "Weekly aggregation of all metrics. Net Sentiment = Positive% - Negative%. Intent Rate = High Intent / Total.",
     "Pandas groupby", "Weekly trends, Peak week, Intent rate"),
]

for i, row in enumerate(methods):
    r   = 3 + i
    bg  = DRED_HEX if i==0 else (LGREY if i%2==0 else WHITE_H)
    fg  = WHITE_H if i==0 else '1A1A1A'
    bold_ = i==0
    ws8.row_dimensions[r].height = 55 if i>0 else 22
    for ci, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=ci, value=val)
        cell.font = Font(bold=bold_, size=9 if i>0 else 10,
                         color=fg, name='Arial')
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='left', vertical='top',
                                    wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        cell.border = Border(top=thin,left=thin,right=thin,bottom=thin)

for col, w in [(1,22),(2,25),(3,55),(4,25),(5,35)]:
    set_col_width(ws8, col, w)


# ─── Set sheet tab colors ─────────────────────────────────────────────────────
tab_colors = {
    "📊 KPI Dashboard"     : "C8102E",
    "🏷️ Brand Awareness"   : "E74C3C",
    "😊 Sentiment"         : "2ECC71",
    "🛒 Purchase Intention": "E67E22",
    "⚡ Engagement"        : "2980B9",
    "📈 Campaign Performance": "8E44AD",
    "📋 Raw Data Sample"   : "7F8C8D",
    "📖 Methodology"       : "1A1A1A",
}
for sheet_name, color in tab_colors.items():
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = color


# ─── Save ─────────────────────────────────────────────────────────────────────
output_path = "/home/claude/KitKat_Brand_Analysis_Report.xlsx"
wb.save(output_path)
print(f"   Excel report saved → {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# FINAL CONSOLE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*65)
print("  KITKAT INSTAGRAM ANALYTICS PIPELINE – COMPLETE")
print("═"*65)
print(f"  📊 Total comments analysed  : {total_comments:,}")
print(f"  👥 Unique users             : {df['user_id'].nunique():,}")
print(f"  📅 Date range               : {df['timestamp'].min().date()} → {df['timestamp'].max().date()}")
print(f"\n  MODULE 1 – BRAND AWARENESS")
print(f"  ─────────────────────────────────────────")
print(f"  Brand awareness rate        : {brand_awareness_pct:.2f}%")
print(f"  Brand-mentioned comments    : {brand_aware_cnt:,}")
print(f"  Top keyword                 : '{word_freq[0][0]}' ({word_freq[0][1]:,} times)")
print(f"  Top rival co-mention        : {list(rival_co.items())[0] if rival_co else 'None'}")
print(f"\n  MODULE 2 – PURCHASE INTENTION")
print(f"  ─────────────────────────────────────────")
print(f"  High Purchase Intent        : {intent_dist.get('High Intent',0):,} ({intent_dist.get('High Intent',0)/total_comments*100:.1f}%)")
print(f"  Medium Intent               : {intent_dist.get('Medium Intent',0):,}")
print(f"  Low Intent                  : {intent_dist.get('Low Intent',0):,}")
print(f"  Negative Intent             : {intent_dist.get('Low Intent (Negative)',0):,}")
print(f"\n  MODULE 3 – CUSTOMER ENGAGEMENT")
print(f"  ─────────────────────────────────────────")
print(f"  Avg engagement score        : {avg_eng:.2f}/100")
print(f"  % with @mention             : {pct_reply:.1f}%")
print(f"  % with emoji                : {pct_emoji:.1f}%")
print(f"  Peak hour (UTC)             : {hourly.loc[hourly['comments'].idxmax(), 'hour']}:00")
print(f"  Most active user            : @{power_users.iloc[0]['username']} ({power_users.iloc[0]['comment_count']} comments)")
print(f"\n  MODULE 4 – CAMPAIGN PERFORMANCE")
print(f"  ─────────────────────────────────────────")
print(f"  % Positive sentiment        : {sent_dist.get('Positive',0)/total_comments*100:.1f}%")
print(f"  % Negative sentiment        : {sent_dist.get('Negative',0)/total_comments*100:.1f}%")
print(f"  Avg polarity                : {df['polarity'].mean():.4f}")
print(f"  Peak activity week          : {peak_week}")
print(f"  Peak week volume            : {peak_vol:,} comments")
print(f"\n  OUTPUT FILES")
print(f"  ─────────────────────────────────────────")
print(f"  📁 Excel Report  : KitKat_Brand_Analysis_Report.xlsx")
print(f"  📊 Charts (5)    : charts/01–05_*.png")
print("═"*65)
